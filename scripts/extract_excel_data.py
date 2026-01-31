import openpyxl
import json

# Load the Excel file
wb = openpyxl.load_workbook('Olkkari_Menu_Wine_Cocktails.xlsx')

print("Available sheets:", wb.sheetnames)

# Extract "Olkkari New Menu" sheet
menu_items = []
if "Olkkari New Menu" in wb.sheetnames:
    sheet = wb["Olkkari New Menu"]
    
    # Get headers from first row
    headers = [cell.value for cell in sheet[1]]
    print(f"\nMenu Headers: {headers}")
    
    # Extract data rows
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if any(row):  # Skip empty rows
            item = {}
            for i, header in enumerate(headers):
                if header and i < len(row):
                    item[header] = row[i]
            if item:
                menu_items.append(item)
    
    print(f"Extracted {len(menu_items)} menu items")
    
    # Save to JSON
    with open('scripts/menu_data.json', 'w', encoding='utf-8') as f:
        json.dump(menu_items, f, indent=2, ensure_ascii=False)
    print("Saved to scripts/menu_data.json")

# Extract "Cocktail list" sheet
cocktail_items = []
if "Cocktail list" in wb.sheetnames:
    sheet = wb["Cocktail list"]
    
    # Get headers from first row
    headers = [cell.value for cell in sheet[1]]
    print(f"\nCocktail Headers: {headers}")
    
    # Extract data rows
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if any(row):  # Skip empty rows
            item = {}
            for i, header in enumerate(headers):
                if header and i < len(row):
                    item[header] = row[i]
            if item:
                cocktail_items.append(item)
    
    print(f"Extracted {len(cocktail_items)} cocktail items")
    
    # Save to JSON
    with open('scripts/cocktail_data.json', 'w', encoding='utf-8') as f:
        json.dump(cocktail_items, f, indent=2, ensure_ascii=False)
    print("Saved to scripts/cocktail_data.json")

wb.close()

print("\n" + "="*80)
print("Sample Menu Item:")
if menu_items:
    print(json.dumps(menu_items[0], indent=2, ensure_ascii=False))

print("\n" + "="*80)
print("Sample Cocktail Item:")
if cocktail_items:
    print(json.dumps(cocktail_items[0], indent=2, ensure_ascii=False))
